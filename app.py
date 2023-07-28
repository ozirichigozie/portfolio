from flask import Flask, request, render_template
import openpyxl
# from markupsafe import escape

app = Flask(__name__)

wb = openpyxl.load_workbook('userdata.xlsx')
ws = wb.active

ws['A1'] = 'Full name'
ws['B1'] = 'Message'
ws['C1'] = 'Email'
ws['D1'] = 'Priority'

home_url = "http://127.0.0.1:5000/"


@app.route('/', methods=["GET", "POST"])
def get_userdata():
    if request.method == "POST":
        Fullname = request.form.get("fullname")
        Message = request.form.get("message")
        Email = request.form.get("email")
        Priority = request.form.get("priority")
        ws.append([Fullname, Message, Email, Priority])
        wb.save('userdata.xlsx')
        return_outputs = {
            "thank_u_msg": f"<h1>Thank you, {Fullname}!</h1><br><i>Click <a href='{home_url}'>here</a> to return to homepage.</p>",
            "render_template": render_template("index.html")
        }
        return return_outputs["thank_u_msg"]
        # return_outputs["sleep"], return_outputs["render_template"]
    return render_template("index.html")


@app.route('/<name>', methods=["GET", "POST"])
def homepage(name):
    print(f"You are welcome, {name}!")
    if request.method == "POST":
        return get_userdata()
    return render_template("index.html")


@app.route("/css/styles.css")
def style_css():
    return app.send_static_file("css/styles.css")


@app.route("/js/scripts.js")
def script_js():
    return app.send_static_file("js/scripts.js")


@app.route("/img/<image_name>")
def images(image_name):
    return app.send_static_file(f"img/{image_name}")


@app.route("/docs/<filename>")
def documents(filename):
    return app.send_static_file(f"docs/{filename}")


if __name__=='__main__':
    app.run()
